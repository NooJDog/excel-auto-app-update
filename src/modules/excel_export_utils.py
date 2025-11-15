from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os, shutil

def create_workbook():
    return Workbook()

def autofit_columns(ws, max_width=60, min_width=8):
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len=0
        for cell in col:
            v=cell.value
            if v is None: continue
            l=len(str(v))
            if l>max_len: max_len=l
        width=int(max_len*1.2)+1
        if width<min_width: width=min_width
        if width>max_width: width=max_width
        ws.column_dimensions[get_column_letter(col_idx)].width=width

def style_header(ws):
    for cell in ws[1]:
        cell.font=Font(bold=True)
        cell.alignment=Alignment(horizontal="center", vertical="center")
        cell.fill=PatternFill("solid", fgColor="FFEFEF")

def color_diff_cell(cell, diff):
    if diff==0:
        cell.fill=PatternFill("solid", fgColor="D7FFD7")
    elif diff>0:
        cell.fill=PatternFill("solid", fgColor="D7E8FF")
    else:
        cell.fill=PatternFill("solid", fgColor="FFD7D7")

def finalize_sheet(ws):
    ws.freeze_panes="A2"
    ws.auto_filter.ref=ws.dimensions

def safe_save_wb(wb, path):
    try:
        wb.save(path)
    except PermissionError:
        base,ext=os.path.splitext(path)
        alt=base+"_locked"+ext
        try:
            wb.save(alt)
        except Exception:
            pass